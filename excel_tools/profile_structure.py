"""excel结构分析工具（最小实现版本）。"""

from collections import OrderedDict
import os
from pathlib import Path
from typing import Any

from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from excel_tools.mcp_instance import excel_mcp
from excel_tools.exceptions import SheetNotFoundError
from excel_tools.core_utils import OPENPYXL_EXTENSIONS, open_workbook, resolve_excel_file_path
from excel_tools.read_backends import Bounds, inspect_workbook, read_range
from dotenv import load_dotenv

# 自动加载当前目录下的 .env
_ = load_dotenv()

# 最大可记录的模式数量，用于防止超大文件产生过多模式。
MAX_PATTERNS: int = int(os.getenv("MAX_PATTERNS", 10))
# 默认扫描最大行数。
DEFAULT_MAX_ROWS: int = int(os.getenv("DEFAULT_MAX_ROWS", 5000))
# 默认扫描最大列数。
DEFAULT_MAX_COLS: int = int(os.getenv("DEFAULT_MAX_COLS", 100))


def _data_type_char(cell: Any) -> str:
    """把单元格值映射成简化类型码。"""
    # 空值直接记为空。
    if cell.value is None:
        return "_"
    # 读取openpyxl内部类型码。
    data_type = cell.data_type
    # 数字类型。
    if data_type == "n":
        return "n"
    # 日期类型。
    if data_type == "d":
        return "d"
    # 公式类型。
    if data_type == "f":
        return "f"
    # 其余都按字符串处理，保证逻辑简单。
    return "s"


def _value_type_char(value: Any) -> str:
    """把只读后端返回的值映射成简化类型码。"""
    if value is None:
        return "_"
    if isinstance(value, bool):
        return "b"
    if isinstance(value, (int, float)):
        return "n"
    if hasattr(value, "isoformat"):
        return "d"
    return "s"


def _build_merged_map(ws: Worksheet, scan_rows: int, scan_cols: int) -> set[tuple[int, int]]:
    """构建扫描范围内“属于合并区域”的坐标集合。"""
    # 使用set方便后续O(1)判断某个坐标是否在合并区。
    merged_cells: set[tuple[int, int]] = set()
    # 遍历工作表里所有合并区域。
    for merged_range in ws.merged_cells.ranges:
        # 如果合并区域左上角已经在扫描范围外，直接跳过。
        if merged_range.min_row > scan_rows or merged_range.min_col > scan_cols:
            continue
        # 计算和扫描边界的交集，避免越界。
        max_row = min(merged_range.max_row, scan_rows)
        max_col = min(merged_range.max_col, scan_cols)
        # 把交集范围内所有坐标加入集合。
        for row in range(merged_range.min_row, max_row + 1):
            for col in range(merged_range.min_col, max_col + 1):
                merged_cells.add((row, col))
    # 返回坐标集合。
    return merged_cells


def _build_merged_regions(ws: Worksheet, scan_rows: int, scan_cols: int) -> list[dict[str, str]]:
    """收集扫描范围内的合并区域信息。"""
    # 用列表保存区域信息，保持输出简单直观。
    regions: list[dict[str, str]] = []
    # 遍历所有合并区域。
    for merged_range in ws.merged_cells.ranges:
        # 超出扫描起点范围时跳过。
        if merged_range.min_row > scan_rows or merged_range.min_col > scan_cols:
            continue
        # 记录区域字符串和锚点坐标。
        regions.append(
            {
                "range": str(merged_range),
                "anchor": ws.cell(merged_range.min_row, merged_range.min_col).coordinate,
            }
        )
    # 返回区域列表。
    return regions


@excel_mcp.tool()
async def excel_profile_structure(
    sheet: str,
    file_path: str = "",
    max_rows_to_scan: int = DEFAULT_MAX_ROWS,
    max_cols_to_scan: int = DEFAULT_MAX_COLS,
) -> dict[str, object]:
    """分析工作表的结构签名（数据类型 + 合并状态）。"""
    async with resolve_excel_file_path(file_path=file_path) as path:
        if path.suffix.lower() not in OPENPYXL_EXTENSIONS:
            return _profile_readonly_backend(
                path=path,
                sheet=sheet,
                max_rows_to_scan=max_rows_to_scan,
                max_cols_to_scan=max_cols_to_scan,
            )

        workbook = open_workbook(path, read_only=False, data_only=False)
        try:
            # 检查sheet是否存在。
            if sheet not in workbook.sheetnames:
                raise SheetNotFoundError(
                    message=f"表单 '{sheet}' 未找到。",
                    suggested_action=f"可用表单: {workbook.sheetnames}",
                    details={"requested_sheet": sheet, "available_sheets": workbook.sheetnames},
                )

            worksheet = workbook[sheet]
            total_rows = worksheet.max_row or 0
            total_cols = worksheet.max_column or 0

            # 空表时直接返回空结果。
            if total_rows == 0 or total_cols == 0:
                return {
                    "sheet": sheet,
                    "scanned_rows": [0, 0],
                    "scanned_cols": [0, 0],
                    "total_rows_in_sheet": total_rows,
                    "total_cols_in_sheet": total_cols,
                    "sampled": False,
                    # "merged_regions": [],
                    "patterns": {},
                    "pattern_row_spans": {},
                    "backend": "openpyxl",
                    "resolved_file_path": str(path),
                }

            scan_rows = min(total_rows, max_rows_to_scan)
            scan_cols = min(total_cols, max_cols_to_scan)
            sampled = scan_rows < total_rows or scan_cols < total_cols
            # 预构建“坐标是否在合并区”索引。
            merged_map = _build_merged_map(worksheet, scan_rows, scan_cols)
            # merged_regions = _build_merged_regions(worksheet, scan_rows, scan_cols)

            pattern_lookup: OrderedDict[str, str] = OrderedDict()
            pattern_data: dict[str, dict[str, str]] = {}
            # 按模式 id 聚合「连续行区间」，避免同一种 P1 在输出里拆开成多条冗余记录。
            pattern_row_spans: dict[str, list[list[int]]] = {}
            current_pattern_id: str | None = None
            current_range_start: int | None = None

            # 遍历扫描窗口内每一行。
            for row_cells in worksheet.iter_rows(min_row=1, max_row=scan_rows, min_col=1, max_col=scan_cols):
                # 当前行号。
                raw_row_idx = row_cells[0].row
                # 理论上row一定有值，这里做保护避免类型检查报错。
                if raw_row_idx is None:
                    continue
                row_idx = int(raw_row_idx)
                # 拼接本行数据类型签名。
                types_signature = "".join(_data_type_char(cell) for cell in row_cells)
                # 拼接本行合并成员签名。
                merged_signature = "".join("M" if (row_idx, cell.column) in merged_map else "_" for cell in row_cells)
                # 组装完整结构key，防止不同合并状态被误归一类。
                pattern_key = f"{types_signature}|{merged_signature}"

                # 已见过该模式时直接复用id。
                if pattern_key in pattern_lookup:
                    pattern_id = pattern_lookup[pattern_key]
                # 超出模式上限时进入OVERFLOW模式。
                elif len(pattern_lookup) >= MAX_PATTERNS:
                    pattern_id = "OVERFLOW"
                    if pattern_id not in pattern_data:
                        pattern_data[pattern_id] = {"types": "?", "merged": "?"}
                # 新模式时创建新id并记录。
                else:
                    pattern_id = f"P{len(pattern_lookup) + 1}"
                    pattern_lookup[pattern_key] = pattern_id
                    pattern_data[pattern_id] = {"types": types_signature, "merged": merged_signature}

                # 模式变化时先关闭上一段连续块。
                if pattern_id != current_pattern_id:
                    if current_pattern_id is not None and current_range_start is not None:
                        # 把刚结束的一段 [起, 止] 挂到对应模式下面（同一模式可能多段，按自上而下顺序追加）。
                        pattern_row_spans.setdefault(current_pattern_id, []).append(
                            [current_range_start, int(row_idx) - 1]
                        )
                    # 开启新的连续块。
                    current_pattern_id = pattern_id
                    current_range_start = row_idx

            # 收尾：把最后一个连续块加入对应模式。
            if current_pattern_id is not None and current_range_start is not None:
                pattern_row_spans.setdefault(current_pattern_id, []).append([current_range_start, scan_rows])

            # 返回结构分析结果。
            return {
                "sheet": sheet,
                "scanned_rows": [1, scan_rows],
                "scanned_cols": [1, scan_cols],
                "total_rows_in_sheet": total_rows,
                "total_cols_in_sheet": total_cols,
                "sampled": sampled,
                "patterns": pattern_data,
                "pattern_row_spans": pattern_row_spans,
                "backend": "openpyxl",
                "resolved_file_path": str(path),
                # "merged_regions": merged_regions,
            }
        finally:
            # 始终关闭工作簿。
            workbook.close()


def _profile_readonly_backend(
    path: Path,
    sheet: str,
    max_rows_to_scan: int,
    max_cols_to_scan: int,
) -> dict[str, object]:
    """用只读后端分析非 openpyxl 格式，如 .xls/.xlsb/.ods。"""
    backend, summaries = inspect_workbook(path)
    summary = next((item for item in summaries if item.name == sheet), None)
    if summary is None:
        available_sheets = [item.name for item in summaries]
        raise SheetNotFoundError(
            message=f"表单 '{sheet}' 未找到。",
            suggested_action=f"可用表单: {available_sheets}",
            details={"requested_sheet": sheet, "available_sheets": available_sheets},
        )

    total_rows = summary.rows or 0
    total_cols = summary.cols or 0
    if total_rows == 0 or total_cols == 0:
        return {
            "sheet": sheet,
            "scanned_rows": [0, 0],
            "scanned_cols": [0, 0],
            "total_rows_in_sheet": total_rows,
            "total_cols_in_sheet": total_cols,
            "sampled": False,
            "patterns": {},
            "pattern_row_spans": {},
            "backend": backend,
            "resolved_file_path": str(path),
        }

    scan_rows = min(total_rows, max_rows_to_scan)
    scan_cols = min(total_cols, max_cols_to_scan)
    sampled = scan_rows < total_rows or scan_cols < total_cols
    result = read_range(
        path,
        sheet,
        Bounds(min_col=1, min_row=1, max_col=scan_cols, max_row=scan_rows),
        analyze_merged_cells=True,
        fill_merged_cells=False,
    )
    merged_map = _merged_cells_to_map(result.merged_cells, scan_rows, scan_cols)

    pattern_lookup: OrderedDict[str, str] = OrderedDict()
    pattern_data: dict[str, dict[str, str]] = {}
    pattern_row_spans: dict[str, list[list[int]]] = {}
    current_pattern_id: str | None = None
    current_range_start: int | None = None

    for offset, row_values in enumerate(result.values):
        row_idx = offset + 1
        types_signature = "".join(_value_type_char(value) for value in row_values)
        merged_signature = "".join(
            "M" if (row_idx, col_idx) in merged_map else "_" for col_idx in range(1, scan_cols + 1)
        )
        pattern_key = f"{types_signature}|{merged_signature}"

        if pattern_key in pattern_lookup:
            pattern_id = pattern_lookup[pattern_key]
        elif len(pattern_lookup) >= MAX_PATTERNS:
            pattern_id = "OVERFLOW"
            if pattern_id not in pattern_data:
                pattern_data[pattern_id] = {"types": "?", "merged": "?"}
        else:
            pattern_id = f"P{len(pattern_lookup) + 1}"
            pattern_lookup[pattern_key] = pattern_id
            pattern_data[pattern_id] = {"types": types_signature, "merged": merged_signature}

        if pattern_id != current_pattern_id:
            if current_pattern_id is not None and current_range_start is not None:
                pattern_row_spans.setdefault(current_pattern_id, []).append([current_range_start, row_idx - 1])
            current_pattern_id = pattern_id
            current_range_start = row_idx

    if current_pattern_id is not None and current_range_start is not None:
        pattern_row_spans.setdefault(current_pattern_id, []).append([current_range_start, scan_rows])

    return {
        "sheet": sheet,
        "scanned_rows": [1, scan_rows],
        "scanned_cols": [1, scan_cols],
        "total_rows_in_sheet": total_rows,
        "total_cols_in_sheet": total_cols,
        "sampled": sampled,
        "patterns": pattern_data,
        "pattern_row_spans": pattern_row_spans,
        "backend": backend,
        "resolved_file_path": str(path),
    }


def _merged_cells_to_map(
    merged_cells: list[dict[str, Any]],
    scan_rows: int,
    scan_cols: int,
) -> set[tuple[int, int]]:
    merged_map: set[tuple[int, int]] = set()
    for item in merged_cells:
        raw_min_col, raw_min_row, raw_max_col, raw_max_row = range_boundaries(str(item["range"]))
        max_row = min(raw_max_row, scan_rows)
        max_col = min(raw_max_col, scan_cols)
        for row in range(raw_min_row, max_row + 1):
            for col in range(raw_min_col, max_col + 1):
                merged_map.add((row, col))
    return merged_map
