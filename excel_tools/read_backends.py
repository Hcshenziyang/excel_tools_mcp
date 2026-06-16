"""Read-only workbook helpers shared by range, inspect, and profile tools."""

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from excel_tools.core_utils import OPENPYXL_EXTENSIONS, open_workbook
from excel_tools.exceptions import FileCorruptedError, SheetNotFoundError, UnsupportedFileError


@dataclass(frozen=True)
class Bounds:
    min_col: int
    min_row: int
    max_col: int
    max_row: int


@dataclass(frozen=True)
class MergedCellRange:
    bounds: Bounds
    anchor_value: Any


@dataclass(frozen=True)
class RangeReadResult:
    backend: str
    values: list[list[Any]]
    merged_cells: list[dict[str, Any]]
    merged_cells_analyzed: bool
    merged_cells_filled: bool


@dataclass(frozen=True)
class SheetSummary:
    name: str
    rows: int | None
    cols: int | None
    dimension_source: str


def normalize_range(start_cell: str, end_cell: str | None) -> tuple[str, str]:
    """Normalize either A1:B2 or separate start/end inputs."""
    if ":" in start_cell and end_cell is None:
        parts = start_cell.split(":")
        return parts[0], parts[1]
    return start_cell, (end_cell or start_cell)


def parse_range(start_cell: str, end_cell: str | None) -> tuple[str, Bounds]:
    """Parse an Excel range and normalize reversed corners."""
    norm_start, norm_end = normalize_range(start_cell, end_cell)
    min_col, min_row, max_col, max_row = range_boundaries(f"{norm_start}:{norm_end}")
    bounds = Bounds(
        min_col=min(min_col, max_col),
        min_row=min(min_row, max_row),
        max_col=max(min_col, max_col),
        max_row=max(min_row, max_row),
    )
    return format_bounds(bounds), bounds


def format_bounds(bounds: Bounds) -> str:
    """Format 1-based bounds as an A1 range."""
    start = f"{get_column_letter(bounds.min_col)}{bounds.min_row}"
    end = f"{get_column_letter(bounds.max_col)}{bounds.max_row}"
    return f"{start}:{end}"


def read_range(
    path: Path,
    sheet: str,
    bounds: Bounds,
    analyze_merged_cells: bool = False,
    fill_merged_cells: bool = False,
) -> RangeReadResult:
    """Read a worksheet range without modifying the source file."""
    if path.suffix.lower() in OPENPYXL_EXTENSIONS:
        return _read_openpyxl_range(path, sheet, bounds, analyze_merged_cells, fill_merged_cells)
    return _read_calamine_range(path, sheet, bounds, analyze_merged_cells, fill_merged_cells)


def inspect_workbook(path: Path) -> tuple[str, list[SheetSummary]]:
    """Return read-only sheet summaries for any supported workbook."""
    if path.suffix.lower() in OPENPYXL_EXTENSIONS:
        return _inspect_openpyxl_workbook(path)
    return _inspect_calamine_workbook(path)


def _read_openpyxl_range(
    path: Path,
    sheet: str,
    bounds: Bounds,
    analyze_merged_cells: bool,
    fill_merged_cells: bool,
) -> RangeReadResult:
    workbook = open_workbook(path, read_only=False, data_only=False)
    try:
        if sheet not in workbook.sheetnames:
            raise SheetNotFoundError(
                message=f"表单 '{sheet}' 未找到。",
                suggested_action=f"可用表单: {workbook.sheetnames}",
                details={"requested_sheet": sheet, "available_sheets": workbook.sheetnames},
            )

        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(
            min_row=bounds.min_row,
            max_row=bounds.max_row,
            min_col=bounds.min_col,
            max_col=bounds.max_col,
        )
        values = [[cell.value for cell in row_cells] for row_cells in rows]
        merged_ranges: list[MergedCellRange] = []

        if analyze_merged_cells:
            for merged_range in worksheet.merged_cells.ranges:
                merged_bounds = Bounds(
                    min_col=merged_range.min_col,
                    min_row=merged_range.min_row,
                    max_col=merged_range.max_col,
                    max_row=merged_range.max_row,
                )
                if not _rects_intersect(bounds, merged_bounds):
                    continue
                anchor_value = worksheet.cell(merged_bounds.min_row, merged_bounds.min_col).value
                merged_ranges.append(MergedCellRange(bounds=merged_bounds, anchor_value=anchor_value))

        return _build_result("openpyxl", values, bounds, merged_ranges, analyze_merged_cells, fill_merged_cells)
    finally:
        workbook.close()


def _read_calamine_range(
    path: Path,
    sheet: str,
    bounds: Bounds,
    analyze_merged_cells: bool,
    fill_merged_cells: bool,
) -> RangeReadResult:
    workbook = _open_calamine_workbook(path)
    try:
        sheet_names = list(workbook.sheet_names)
        if sheet not in sheet_names:
            raise SheetNotFoundError(
                message=f"表单 '{sheet}' 未找到。",
                suggested_action=f"可用表单: {sheet_names}",
                details={"requested_sheet": sheet, "available_sheets": sheet_names},
            )

        worksheet = workbook.get_sheet_by_name(sheet)
        rows = worksheet.to_python(skip_empty_area=False, nrows=bounds.max_row)
        values = _slice_rows(rows, bounds)
        merged_ranges: list[MergedCellRange] = []

        if analyze_merged_cells:
            merged_ranges = _calamine_merged_ranges(worksheet, rows, bounds)

        return _build_result("python-calamine", values, bounds, merged_ranges, analyze_merged_cells, fill_merged_cells)
    finally:
        workbook.close()


def _inspect_openpyxl_workbook(path: Path) -> tuple[str, list[SheetSummary]]:
    workbook = open_workbook(path, read_only=False, data_only=False)
    try:
        summaries = [
            SheetSummary(
                name=sheet_name,
                rows=workbook[sheet_name].max_row,
                cols=workbook[sheet_name].max_column,
                dimension_source="declared",
            )
            for sheet_name in workbook.sheetnames
        ]
        return "openpyxl", summaries
    finally:
        workbook.close()


def _inspect_calamine_workbook(path: Path) -> tuple[str, list[SheetSummary]]:
    workbook = _open_calamine_workbook(path)
    try:
        summaries: list[SheetSummary] = []
        for sheet_name in workbook.sheet_names:
            worksheet = workbook.get_sheet_by_name(sheet_name)
            if worksheet.end is None:
                rows, cols = 0, 0
            else:
                rows, cols = worksheet.end[0] + 1, worksheet.end[1] + 1
            summaries.append(
                SheetSummary(
                    name=sheet_name,
                    rows=rows,
                    cols=cols,
                    dimension_source="calamine",
                )
            )
        return "python-calamine", summaries
    finally:
        workbook.close()


def _open_calamine_workbook(path: Path):
    try:
        from python_calamine import CalamineWorkbook
    except ImportError as e:
        raise UnsupportedFileError(
            message="读取此格式需要安装 python-calamine。",
            suggested_action="安装 python-calamine，或将文件转换为.xlsx后再处理。",
            details={"path": str(path), "extension": path.suffix},
        ) from e

    try:
        return CalamineWorkbook.from_path(path)
    except Exception as e:
        msg = str(e).lower()
        if "unsupported" in msg or "format" in msg:
            raise UnsupportedFileError(
                message=f"不支持的Excel文件格式: {e}",
                suggested_action="请确认文件扩展名和真实格式一致，或转换为.xlsx。",
                details={"path": str(path), "extension": path.suffix},
            ) from e
        raise FileCorruptedError(
            message=f"无法打开文件: {e}",
            suggested_action="验证文件是否损坏或加密。",
            details={"path": str(path), "underlying_error": type(e).__name__},
        ) from e


def _slice_rows(rows: list[list[Any]], bounds: Bounds) -> list[list[Any]]:
    values: list[list[Any]] = []
    width = bounds.max_col - bounds.min_col + 1
    for row_idx in range(bounds.min_row - 1, bounds.max_row):
        if row_idx >= len(rows):
            values.append([None] * width)
            continue
        source_row = rows[row_idx]
        row_values: list[Any] = []
        for col_idx in range(bounds.min_col - 1, bounds.max_col):
            if col_idx >= len(source_row):
                row_values.append(None)
            else:
                value = source_row[col_idx]
                row_values.append(None if value == "" else value)
        values.append(row_values)
    return values


def _calamine_merged_ranges(worksheet, rows: list[list[Any]], requested: Bounds) -> list[MergedCellRange]:
    merged_ranges: list[MergedCellRange] = []
    for start, end in worksheet.merged_cell_ranges:
        merged_bounds = Bounds(
            min_row=start[0] + 1,
            min_col=start[1] + 1,
            max_row=end[0] + 1,
            max_col=end[1] + 1,
        )
        if not _rects_intersect(requested, merged_bounds):
            continue
        anchor_value = _value_from_rows(rows, merged_bounds.min_row, merged_bounds.min_col)
        merged_ranges.append(MergedCellRange(bounds=merged_bounds, anchor_value=anchor_value))
    return merged_ranges


def _value_from_rows(rows: list[list[Any]], row: int, col: int) -> Any:
    row_idx = row - 1
    col_idx = col - 1
    if row_idx >= len(rows) or col_idx >= len(rows[row_idx]):
        return None
    value = rows[row_idx][col_idx]
    return None if value == "" else value


def _build_result(
    backend: str,
    values: list[list[Any]],
    requested: Bounds,
    merged_ranges: list[MergedCellRange],
    analyze_merged_cells: bool,
    fill_merged_cells: bool,
) -> RangeReadResult:
    merged_cells: list[dict[str, Any]] = []
    if analyze_merged_cells:
        for merged_range in merged_ranges:
            if fill_merged_cells:
                _fill_merged_values(values, requested, merged_range)
            merged_cells.append(
                {
                    "range": format_bounds(merged_range.bounds),
                    "anchor": f"{get_column_letter(merged_range.bounds.min_col)}{merged_range.bounds.min_row}",
                    "value": merged_range.anchor_value,
                    "applied": fill_merged_cells,
                }
            )

    return RangeReadResult(
        backend=backend,
        values=values,
        merged_cells=merged_cells,
        merged_cells_analyzed=analyze_merged_cells,
        merged_cells_filled=analyze_merged_cells and fill_merged_cells,
    )


def _fill_merged_values(
    values: list[list[Any]],
    requested: Bounds,
    merged_range: MergedCellRange,
) -> None:
    fill_bounds = _intersection(requested, merged_range.bounds)
    if fill_bounds is None:
        return
    for row in range(fill_bounds.min_row, fill_bounds.max_row + 1):
        for col in range(fill_bounds.min_col, fill_bounds.max_col + 1):
            values[row - requested.min_row][col - requested.min_col] = merged_range.anchor_value


def _intersection(left: Bounds, right: Bounds) -> Bounds | None:
    min_col = max(left.min_col, right.min_col)
    min_row = max(left.min_row, right.min_row)
    max_col = min(left.max_col, right.max_col)
    max_row = min(left.max_row, right.max_row)
    if min_col > max_col or min_row > max_row:
        return None
    return Bounds(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def _rects_intersect(left: Bounds, right: Bounds) -> bool:
    return _intersection(left, right) is not None
