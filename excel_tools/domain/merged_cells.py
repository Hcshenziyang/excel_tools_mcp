"""Merged-cell normalization helpers."""

from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries

from excel_tools.domain.ranges import format_bounds, intersection
from excel_tools.models.schemas import Bounds, MergedCellInfo, MergedCellRange, RangeReadResult


def build_range_read_result(
    *,
    backend: str,
    values: list[list[Any]],
    requested: Bounds,
    merged_ranges: list[MergedCellRange],
    analyze_merged_cells: bool,
    fill_merged_cells: bool,
) -> RangeReadResult:
    """Build a unified range-read result and optionally fill merged cells."""
    merged_cells: list[MergedCellInfo] = []
    if analyze_merged_cells:
        for merged_range in merged_ranges:
            if fill_merged_cells:
                fill_merged_values(values, requested, merged_range)
            merged_cells.append(
                MergedCellInfo(
                    range=format_bounds(merged_range.bounds),
                    anchor=f"{get_column_letter(merged_range.bounds.min_col)}{merged_range.bounds.min_row}",
                    value=merged_range.anchor_value,
                    applied=fill_merged_cells,
                )
            )

    return RangeReadResult(
        backend=backend,
        values=values,
        merged_cells=merged_cells,
        merged_cells_analyzed=analyze_merged_cells,
        merged_cells_filled=analyze_merged_cells and fill_merged_cells,
    )


def fill_merged_values(
    values: list[list[Any]],
    requested: Bounds,
    merged_range: MergedCellRange,
) -> None:
    """Fill the requested slice of a merged range with its anchor value."""
    fill_bounds = intersection(requested, merged_range.bounds)
    if fill_bounds is None:
        return
    for row in range(fill_bounds.min_row, fill_bounds.max_row + 1):
        for col in range(fill_bounds.min_col, fill_bounds.max_col + 1):
            values[row - requested.min_row][col - requested.min_col] = merged_range.anchor_value


def merged_cells_to_map(
    merged_cells: list[MergedCellInfo],
    scan_rows: int,
    scan_cols: int,
) -> set[tuple[int, int]]:
    """Convert merged-cell metadata to a coordinate set clipped to a scan window."""
    merged_map: set[tuple[int, int]] = set()
    for item in merged_cells:
        raw_min_col, raw_min_row, raw_max_col, raw_max_row = range_boundaries(item.range)
        if raw_min_col is None or raw_min_row is None or raw_max_col is None or raw_max_row is None:
            continue
        max_row = min(raw_max_row, scan_rows)
        max_col = min(raw_max_col, scan_cols)
        for row in range(raw_min_row, max_row + 1):
            for col in range(raw_min_col, max_col + 1):
                merged_map.add((row, col))
    return merged_map
