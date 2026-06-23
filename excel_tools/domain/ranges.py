"""Excel range parsing and rectangle helpers."""

from openpyxl.utils import get_column_letter, range_boundaries

from excel_tools.exceptions import InvalidParameterError
from excel_tools.models.schemas import Bounds


def format_bounds(bounds: Bounds) -> str:
    """Format 1-based bounds as an A1 range."""
    start = f"{get_column_letter(bounds.min_col)}{bounds.min_row}"
    end = f"{get_column_letter(bounds.max_col)}{bounds.max_row}"
    return f"{start}:{end}"


def parse_range(start_cell: str, end_cell: str | None) -> tuple[str, Bounds]:
    """返回一个元组，包含格式化的范围字符串和对应的 Bounds 对象。
    例如，输入 "A1" 和 "C3" 将返回
    ("A1:C3", Bounds(min_col=1, min_row=1, max_col=3, max_row=3))。
    """
    if ":" in start_cell and end_cell is None:
        parts = start_cell.split(":")
        norm_start, norm_end = parts[0], parts[1]
    else:
        norm_start, norm_end = start_cell, (end_cell or start_cell)

    min_col, min_row, max_col, max_row = range_boundaries(f"{norm_start}:{norm_end}")
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise InvalidParameterError(
            message=f"无效的单元格范围: {norm_start}:{norm_end}",
            suggested_action="请使用完整单元格坐标范围，例如 A1:D20，或单个单元格 A1。",
            details={"start_cell": start_cell, "end_cell": end_cell},
        )

    bounds = Bounds(
        min_col=min(min_col, max_col),
        min_row=min(min_row, max_row),
        max_col=max(min_col, max_col),
        max_row=max(min_row, max_row),
    )
    return format_bounds(bounds), bounds


def intersection(left: Bounds, right: Bounds) -> Bounds | None:
    """Return the intersection of two 1-based rectangular bounds."""
    min_col = max(left.min_col, right.min_col)
    min_row = max(left.min_row, right.min_row)
    max_col = min(left.max_col, right.max_col)
    max_row = min(left.max_row, right.max_row)
    if min_col > max_col or min_row > max_row:
        return None
    return Bounds(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def rects_intersect(left: Bounds, right: Bounds) -> bool:
    """Return whether two 1-based rectangular bounds intersect."""
    return intersection(left, right) is not None
