"""Unmerge worksheet cells."""

from typing import cast

from openpyxl.utils import range_boundaries

from excel_tools.exceptions import InvalidParameterError, SheetNotFoundError
from excel_tools.models.schemas import UnmergeCellsResult
from excel_tools.readers.files import open_workbook, resolve_excel_file_path


def _normalize_bounds(min_col: int, min_row: int, max_col: int, max_row: int) -> tuple[int, int, int, int]:
    lc, rc = min(min_col, max_col), max(min_col, max_col)
    lr, rr = min(min_row, max_row), max(min_row, max_row)
    return lc, lr, rc, rr


def _rects_intersect(
    ax1: int,
    ay1: int,
    ax2: int,
    ay2: int,
    bx1: int,
    by1: int,
    bx2: int,
    by2: int,
) -> bool:
    return not (ax2 < bx1 or ax1 > bx2 or ay2 < by1 or ay1 > by2)


async def unmerge_cells(
    *,
    file_path: str,
    sheet: str,
    start_cell: str,
    end_cell: str,
    fill_with_anchor_value: bool,
) -> UnmergeCellsResult:
    """Unmerge ranges intersecting a requested rectangle and optionally fill anchor values."""
    async with resolve_excel_file_path(file_path=file_path) as path:
        target_range = f"{start_cell}:{end_cell}"
        raw_min_col, raw_min_row, raw_max_col, raw_max_row = cast(
            tuple[int | None, int | None, int | None, int | None],
            range_boundaries(target_range),
        )
        if raw_min_col is None or raw_min_row is None or raw_max_col is None or raw_max_row is None:
            raise InvalidParameterError(
                message=f"无效的单元格范围: {target_range}",
                suggested_action="请使用完整单元格坐标范围，例如 A1:D20。",
                details={"start_cell": start_cell, "end_cell": end_cell},
            )
        req_min_col, req_min_row, req_max_col, req_max_row = _normalize_bounds(
            raw_min_col, raw_min_row, raw_max_col, raw_max_row
        )

        workbook = open_workbook(path, read_only=False, data_only=False)
        try:
            if sheet not in workbook.sheetnames:
                raise SheetNotFoundError(
                    message=f"表单 '{sheet}' 未找到。",
                    suggested_action=f"可用表单: {workbook.sheetnames}",
                    details={"requested_sheet": sheet, "available_sheets": workbook.sheetnames},
                )

            worksheet = workbook[sheet]
            unmerged_ranges: list[str] = []

            for merged_range in list(worksheet.merged_cells.ranges):
                if not _rects_intersect(
                    req_min_col,
                    req_min_row,
                    req_max_col,
                    req_max_row,
                    merged_range.min_col,
                    merged_range.min_row,
                    merged_range.max_col,
                    merged_range.max_row,
                ):
                    continue
                anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
                worksheet.unmerge_cells(str(merged_range))
                if fill_with_anchor_value:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            worksheet.cell(row=row, column=col, value=anchor_value)
                unmerged_ranges.append(str(merged_range))

            workbook.save(path)
            return UnmergeCellsResult(
                sheet=sheet,
                requested_range=target_range,
                unmerged_count=len(unmerged_ranges),
                unmerged_ranges=unmerged_ranges,
                fill_with_anchor_value=fill_with_anchor_value,
                resolved_file_path=str(path),
            )
        finally:
            workbook.close()
