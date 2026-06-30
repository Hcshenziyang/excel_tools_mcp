"""Analyze worksheet structure signatures."""

from collections import OrderedDict
import os
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl.worksheet.worksheet import Worksheet

from excel_tools.domain.merged_cells import merged_cells_to_map
from excel_tools.domain.sheets import resolve_sheet_name
from excel_tools.models.schemas import Bounds, PatternInfo, ProfileStructureResult
from excel_tools.readers.files import OPENPYXL_EXTENSIONS, open_workbook, resolve_excel_file_path
from excel_tools.readers.selector import get_reader

_ = load_dotenv()

MAX_PATTERNS: int = int(os.getenv("MAX_PATTERNS", 10))
DEFAULT_MAX_ROWS: int = int(os.getenv("DEFAULT_MAX_ROWS", 5000))
DEFAULT_MAX_COLS: int = int(os.getenv("DEFAULT_MAX_COLS", 100))


def _data_type_char(cell: Any) -> str:
    if cell.value is None:
        return "_"
    if cell.data_type == "n":
        return "n"
    if cell.data_type == "d":
        return "d"
    if cell.data_type == "f":
        return "f"
    return "s"


def _value_type_char(value: Any) -> str:
    if value is None:
        return "_"
    if isinstance(value, bool):
        return "b"
    if isinstance(value, (int, float)):
        return "n"
    if hasattr(value, "isoformat"):
        return "d"
    return "s"


def _build_merged_map(ws: Worksheet, scan_rows: int, scan_cols: int) -> set[tuple[int, int]]:
    merged_cells: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row > scan_rows or merged_range.min_col > scan_cols:
            continue
        max_row = min(merged_range.max_row, scan_rows)
        max_col = min(merged_range.max_col, scan_cols)
        for row in range(merged_range.min_row, max_row + 1):
            for col in range(merged_range.min_col, max_col + 1):
                merged_cells.add((row, col))
    return merged_cells


def _build_profile_result(
    *,
    sheet: str,
    scanned_rows: tuple[int, int],
    scanned_cols: tuple[int, int],
    total_rows_in_sheet: int,
    total_cols_in_sheet: int,
    sampled: bool,
    patterns: dict[str, PatternInfo],
    pattern_row_spans: dict[str, list[tuple[int, int]]],
    backend: str,
    resolved_file_path: str,
) -> ProfileStructureResult:
    return ProfileStructureResult(
        sheet=sheet,
        scanned_rows=scanned_rows,
        scanned_cols=scanned_cols,
        total_rows_in_sheet=total_rows_in_sheet,
        total_cols_in_sheet=total_cols_in_sheet,
        sampled=sampled,
        patterns=patterns,
        pattern_row_spans=pattern_row_spans,
        backend=backend,
        resolved_file_path=resolved_file_path,
    )


async def profile_structure(
    *,
    file_path: str,
    sheet: str,
    max_rows_to_scan: int = DEFAULT_MAX_ROWS,
    max_cols_to_scan: int = DEFAULT_MAX_COLS,
) -> ProfileStructureResult:
    """Analyze worksheet row structure signatures."""
    async with resolve_excel_file_path(file_path=file_path) as path:
        reader = get_reader(path)
        if path.suffix.lower() not in OPENPYXL_EXTENSIONS:
            return _profile_reader_backend(
                path=path,
                sheet=sheet,
                max_rows_to_scan=max_rows_to_scan,
                max_cols_to_scan=max_cols_to_scan,
                backend=reader.backend_name,
            )

        workbook = open_workbook(path, read_only=False, data_only=False)
        try:
            resolved_sheet = resolve_sheet_name(sheet, workbook.sheetnames)
            worksheet = workbook[resolved_sheet]
            total_rows = worksheet.max_row or 0
            total_cols = worksheet.max_column or 0

            if total_rows == 0 or total_cols == 0:
                return _build_profile_result(
                    sheet=resolved_sheet,
                    scanned_rows=(0, 0),
                    scanned_cols=(0, 0),
                    total_rows_in_sheet=total_rows,
                    total_cols_in_sheet=total_cols,
                    sampled=False,
                    patterns={},
                    pattern_row_spans={},
                    backend=reader.backend_name,
                    resolved_file_path=str(path),
                )

            scan_rows = min(total_rows, max_rows_to_scan)
            scan_cols = min(total_cols, max_cols_to_scan)
            sampled = scan_rows < total_rows or scan_cols < total_cols
            merged_map = _build_merged_map(worksheet, scan_rows, scan_cols)
            patterns, pattern_row_spans = _profile_rows(
                row_values_iter=(
                    [_data_type_char(cell) for cell in row_cells]
                    for row_cells in worksheet.iter_rows(min_row=1, max_row=scan_rows, min_col=1, max_col=scan_cols)
                ),
                merged_signature_iter=(
                    "".join("M" if (row_idx, col_idx) in merged_map else "_" for col_idx in range(1, scan_cols + 1))
                    for row_idx in range(1, scan_rows + 1)
                ),
            )

            return _build_profile_result(
                sheet=resolved_sheet,
                scanned_rows=(1, scan_rows),
                scanned_cols=(1, scan_cols),
                total_rows_in_sheet=total_rows,
                total_cols_in_sheet=total_cols,
                sampled=sampled,
                patterns=patterns,
                pattern_row_spans=pattern_row_spans,
                backend=reader.backend_name,
                resolved_file_path=str(path),
            )
        finally:
            workbook.close()


def _profile_reader_backend(
    *,
    path: Path,
    sheet: str,
    max_rows_to_scan: int,
    max_cols_to_scan: int,
    backend: str,
) -> ProfileStructureResult:
    reader = get_reader(path)
    summaries = reader.inspect_workbook(path)
    available_sheets = [item.name for item in summaries]
    resolved_sheet = resolve_sheet_name(sheet, available_sheets)
    summary = next(item for item in summaries if item.name == resolved_sheet)

    total_rows = summary.rows or 0
    total_cols = summary.cols or 0
    if total_rows == 0 or total_cols == 0:
        return _build_profile_result(
            sheet=resolved_sheet,
            scanned_rows=(0, 0),
            scanned_cols=(0, 0),
            total_rows_in_sheet=total_rows,
            total_cols_in_sheet=total_cols,
            sampled=False,
            patterns={},
            pattern_row_spans={},
            backend=backend,
            resolved_file_path=str(path),
        )

    scan_rows = min(total_rows, max_rows_to_scan)
    scan_cols = min(total_cols, max_cols_to_scan)
    sampled = scan_rows < total_rows or scan_cols < total_cols
    result = reader.read_range(
        path,
        resolved_sheet,
        Bounds(min_col=1, min_row=1, max_col=scan_cols, max_row=scan_rows),
        analyze_merged_cells=True,
        fill_merged_cells=False,
    )
    merged_map = merged_cells_to_map(result.merged_cells, scan_rows, scan_cols)
    patterns, pattern_row_spans = _profile_rows(
        row_values_iter=([_value_type_char(value) for value in row_values] for row_values in result.values),
        merged_signature_iter=(
            "".join("M" if (row_idx, col_idx) in merged_map else "_" for col_idx in range(1, scan_cols + 1))
            for row_idx in range(1, scan_rows + 1)
        ),
    )

    return _build_profile_result(
        sheet=resolved_sheet,
        scanned_rows=(1, scan_rows),
        scanned_cols=(1, scan_cols),
        total_rows_in_sheet=total_rows,
        total_cols_in_sheet=total_cols,
        sampled=sampled,
        patterns=patterns,
        pattern_row_spans=pattern_row_spans,
        backend=backend,
        resolved_file_path=str(path),
    )


def _profile_rows(
    *,
    row_values_iter,
    merged_signature_iter,
) -> tuple[dict[str, PatternInfo], dict[str, list[tuple[int, int]]]]:
    pattern_lookup: OrderedDict[str, str] = OrderedDict()
    pattern_data: dict[str, PatternInfo] = {}
    pattern_row_spans: dict[str, list[tuple[int, int]]] = {}
    current_pattern_id: str | None = None
    current_range_start: int | None = None
    last_row_idx = 0

    for offset, (type_values, merged_signature) in enumerate(zip(row_values_iter, merged_signature_iter)):
        row_idx = offset + 1
        last_row_idx = row_idx
        types_signature = "".join(type_values)
        pattern_key = f"{types_signature}|{merged_signature}"

        if pattern_key in pattern_lookup:
            pattern_id = pattern_lookup[pattern_key]
        elif len(pattern_lookup) >= MAX_PATTERNS:
            pattern_id = "OVERFLOW"
            if pattern_id not in pattern_data:
                pattern_data[pattern_id] = PatternInfo(types="?", merged="?")
        else:
            pattern_id = f"P{len(pattern_lookup) + 1}"
            pattern_lookup[pattern_key] = pattern_id
            pattern_data[pattern_id] = PatternInfo(types=types_signature, merged=merged_signature)

        if pattern_id != current_pattern_id:
            if current_pattern_id is not None and current_range_start is not None:
                pattern_row_spans.setdefault(current_pattern_id, []).append((current_range_start, row_idx - 1))
            current_pattern_id = pattern_id
            current_range_start = row_idx

    if current_pattern_id is not None and current_range_start is not None:
        pattern_row_spans.setdefault(current_pattern_id, []).append((current_range_start, last_row_idx))

    return pattern_data, pattern_row_spans
