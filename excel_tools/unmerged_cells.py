# -*- coding:utf-8 -*-
"""
Excel 文件解除合并单元格并填充数据工具
"""

from excel_tools.mcp_instance import excel_mcp

from typing import cast

from excel_tools.exceptions import SheetNotFoundError
from excel_tools.core_utils import open_workbook, resolve_excel_file_path
from openpyxl.utils import range_boundaries


def _normalize_bounds(min_col: int, min_row: int, max_col: int, max_row: int) -> tuple[int, int, int, int]:
    """把 range_boundaries 的结果规范成真实的 min/max（颠倒角点时 openpyxl 不会自动交换）。"""
    lc, rc = min(min_col, max_col), max(min_col, max_col)
    lr, rr = min(min_row, max_row), max(min_row, max_row)
    return lc, lr, rc, rr


def _rects_intersect(
    ax1: int,
    ay1: int,
    ax2: int,
    ay2: int,
    bx1: int,
    by1: int,
    bx2: int,
    by2: int,
) -> bool:
    """两轴对齐矩形是否相交（含边相邻）。"""
    return not (ax2 < bx1 or ax1 > bx2 or ay2 < by1 or ay1 > by2)


# 暂不注册为 MCP 工具：真实解除合并会修改源文件，默认工具面先改为只读归一化读取。
# @excel_mcp.tool()
async def excel_unmerge_cells(
    sheet: str,
    start_cell: str,
    end_cell: str,
    file_path: str = "",
    fill_with_anchor_value: bool = True,
) -> dict[str, object]:
    """解除与输入矩形相交的合并区整段（不可只拆一部分），并可选用锚点值填满原合并格。"""
    async with resolve_excel_file_path(file_path=file_path) as path:
        # 组合目标范围（仅用于返回展示）。
        target_range = f"{start_cell}:{end_cell}"
        # 解析用户框选矩形并规范化（合法 A1:D6 式串会得到四个 int）。
        raw_min_col, raw_min_row, raw_max_col, raw_max_row = cast(
            tuple[int, int, int, int],
            range_boundaries(target_range),
        )
        req_min_col, req_min_row, req_max_col, req_max_row = _normalize_bounds(
            raw_min_col, raw_min_row, raw_max_col, raw_max_row
        )
        # 打开工作簿用于修改。
        workbook = open_workbook(path, read_only=False, data_only=False)
        try:
            # 检查表名是否存在。
            if sheet not in workbook.sheetnames:
                raise SheetNotFoundError(
                    message=f"表单 '{sheet}' 未找到。",
                    suggested_action=f"可用表单: {workbook.sheetnames}",
                    details={"requested_sheet": sheet, "available_sheets": workbook.sheetnames},
                )

            # 获取工作表。
            worksheet = workbook[sheet]
            # 记录成功解除的范围，便于返回。
            unmerged_ranges: list[str] = []

            # 遍历当前所有合并范围副本，避免边遍历边修改引发问题。
            for merged_range in list(worksheet.merged_cells.ranges):
                # 与用户输入区域有交集则整段解除（与 Excel 常见行为一致：选中区内触及的合并会整片拆开）。
                if not _rects_intersect(
                    req_min_col,
                    req_min_row,
                    req_max_col,
                    req_max_row,
                    merged_range.min_col,
                    merged_range.min_row,
                    merged_range.max_col,
                    merged_range.max_row,
                ):
                    continue
                # 读取锚点单元格值（左上角）。
                anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
                # 执行解除合并。
                worksheet.unmerge_cells(str(merged_range))
                # 需要回填时，把整个原合并区填成锚点值。
                if fill_with_anchor_value:
                    for row in range(merged_range.min_row, merged_range.max_row + 1):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            worksheet.cell(row=row, column=col, value=anchor_value)
                # 记录处理结果。
                unmerged_ranges.append(str(merged_range))

            # 保存文件。
            workbook.save(path)
            # 返回结果。
            return {
                "sheet": sheet,
                "requested_range": target_range,
                "unmerged_count": len(unmerged_ranges),
                "unmerged_ranges": unmerged_ranges,
                "fill_with_anchor_value": fill_with_anchor_value,
                "resolved_file_path": str(path),
            }
        finally:
            # 关闭工作簿。
            workbook.close()

